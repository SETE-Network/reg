from django.shortcuts import render
from django.views.generic import TemplateView
import requests
from glob import glob
import pandas as pd
import os
import wget
from django.core.files.storage import FileSystemStorage
import time
import webbrowser
import xlsxwriter
import openpyxl
from openpyxl import load_workbook
import openpyxl as xl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.table import Table, TableStyleInfo





class Home(TemplateView):
    template_name = "home.html"

def upload(request):
    for f in glob('./static/*.csv'):
      os.remove(f)
    if request.method == 'POST':
        uploaded_file = request.FILES['document']
        fs = FileSystemStorage()
        fs.save(uploaded_file.name, uploaded_file)
    return render(request, 'upload.html')

def button(request):

    return render(request,'home.html')

def output(request):

    for f in glob('./static/*.csv'):
        df = pd.read_csv(f)
    df.to_excel('./static/working.xlsx', index=False)

    df = pd.read_excel('./static/working.xlsx')


    df['disucssion total'] = df.loc[:, (df.columns.str.startswith('Discussion 1')) +
                                (df.columns.str.startswith('Discussion 2')) +
                                (df.columns.str.startswith('Discussion 3')) +
                                (df.columns.str.startswith('Discussion 4')) +
                                (df.columns.str.startswith('Discussion 5'))].mean(axis=1)
    #                           ENG103
    selected_column = df.loc[:, (df.columns.str.startswith('Username')) +
                                (df.columns.str.contains('disucssion total')) +
                                (df.columns.str.contains('Research Introduction and background')) +
                                (df.columns.str.contains('Research Finding')) +
                                (df.columns.str.contains('Research Paper')) +
                                (df.columns.str.contains('Research Presentation')) +
                                (df.columns.str.contains('articipati')) +
    #                           ENG102
                                (df.columns.str.contains('Expository Essay')) +
                                (df.columns.str.contains('Annotated Bibliography')) +
                                (df.columns.str.contains('Causal Analysis Essay')) +
    #                           ENG101
                                (df.columns.str.contains('Midterm')) +
                                (df.columns.str.contains('Summary Response')) +
                                (df.columns.str.contains('Argumentative Essay')) +
                                (df.columns.str.contains('Final Exam')) +
                                (df.columns.str.contains('Presntation')) +
                                (df.columns.str.contains('Presentation')) +
                                (df.columns.str.contains('PRESENTATION'))]


    new_df = selected_column.copy()

        #new_df[new_df.columns[5]] = selected_column.copy()
    new_df.insert(2, 'Username2', new_df['Username'])
    new_df.insert(4, 'Username3', new_df['Username'])
    new_df.insert(6, 'Username4', new_df['Username'])
    new_df.insert(8, 'Username5', new_df['Username'])
    new_df.insert(10, 'Username6', new_df['Username'])
    new_df.insert(12, 'Username7', new_df['Username'])
    print(new_df)

    # new_df.to_csv('output/rg-ready.xlsx', index=False)

    new_df.to_excel('./assets/rg-ready.xlsx', index=False)

    # Create new Sheet
    wb2 = load_workbook('./assets/rg-ready.xlsx')
    wb2.create_sheet('FAT Report')
    wb2.save('./assets/rg-ready.xlsx')


    # df = pd.read_excel('./assets/rg-ready.xlsx')

    # importing openpyxl module


    # opening the source excel file
    filename ="./static/working.xlsx"
    wb1 = xl.load_workbook(filename)
    ws1 = wb1.worksheets[0]

    # opening the destination excel file
    filename1 ="./assets/rg-ready.xlsx"
    wb2 = xl.load_workbook(filename1)
    ws2 = wb2.worksheets[1]

    # calculate total number of rows and
    # columns in source excel file
    mr = ws1.max_row
    mc = ws1.max_column

    # copying the cell values from source
    # excel file to destination excel file
    for i in range (1, mr + 1):
        for j in range (1, mc + 1):
            # reading cell value from source excel file
            c = ws1.cell(row = i, column = j)

            # writing the read value to destination excel file
            ws2.cell(row = i, column = j).value = c.value

    # saving the destination excel file
    wb2.save(str(filename1))



    # wb.save('upload/working2.xlsx')





    # df2 = df1.copy()
    # with pd.ExcelWriter('output/rg-ready.xlsx') as writer:
    #     df1.to_excel(writer, sheet_name='sheet1')
    #     df2.to_excel(writer, sheet_name='FAT Report')


    # for f in glob('upload/*.csv'):
    #     os.remove(f)

    df = pd.read_excel('./static/working.xlsx')
    selected_column = df.loc[:, (df.columns.str.startswith('Username')) +
                                (df.columns.str.startswith('Last Name')) +
                                (df.columns.str.startswith('First Name')) +
                                (df.columns.str.startswith('Running Total')) +
                                (df.columns.str.startswith('Discussion 1')) +
                                (df.columns.str.startswith('Discussion 2')) +
                                (df.columns.str.startswith('Discussion 3')) +
                                (df.columns.str.startswith('Discussion 4')) +
                                (df.columns.str.startswith('Discussion 5')) +
    #                           ENG103
                                (df.columns.str.contains('Research Introduction and background')) +
                                (df.columns.str.contains('Research Finding')) +
                                (df.columns.str.contains('Research Paper')) +
                                (df.columns.str.contains('Research Presentation')) +
                                (df.columns.str.contains('articipati')) +
    #                           ENG102
                                (df.columns.str.contains('Expository Essay')) +
                                (df.columns.str.contains('Annotated Bibliography')) +
                                (df.columns.str.contains('Causal Analysis Essay')) +
    #                           ENG101
                                (df.columns.str.contains('Midterm')) +
                                (df.columns.str.contains('Summary Response')) +
                                (df.columns.str.contains('Argumentative Essay')) +
                                (df.columns.str.contains('Final Exam')) +
                                (df.columns.str.contains('Presntation')) +
                                (df.columns.str.contains('Presentation')) +
                                (df.columns.str.contains('PRESENTATION'))]
    df2 = selected_column.copy()

    df3 = pd.read_excel('./assets/rg-ready.xlsx')
    selected_column = df3
    path2 = './assets/rg-ready.xlsx'

    writer = pd.ExcelWriter(path2, engine = 'xlsxwriter')
    df2.to_excel(writer, sheet_name = 'FAT Report', index=False)
    df3.to_excel(writer, sheet_name = 'Reg-Ready', index=False)
    writer.save()









    # df2.to_excel("output/rg-ready.xlsx",
    #              sheet_name='FAT Report', index=False, overwrite=False)


    # i = 1
    # while i <= ws.max_row:
    #     print(f'i = {i}\tcell value (i, 1) is {ws.cell(row=1, column=i).value}')
    #     if ws.cell(row=1, column=i).value in ['Student ID','Last Access','Availability','Nickname', 'Total', '', '', '']:
    #         ws.delete_cols(i, 1)
    #     else:
    #         i += 1
    # ws.insert_rows(1, 4,)
    # ws['A2'] = 1
    # wb.save('output/rg-ready.xlsx')

    wb = openpyxl.load_workbook('./assets/rg-ready.xlsx')
    wb.sheetnames
    ws = wb['FAT Report']

    # Styles
    br = border = Border(bottom=Side(border_style='thin', color='000000'))



    ws.insert_rows(0, 3)
    ws.merge_cells('c1:l1')
    ws['c1'].value = "FAT Report"
    a1 = ws['b1']
    ws['a1'].font = Font(size = "20")
    ws['b2'].font = Font(size = "16")
    ws['d2'].font = Font(size = "16")

    ws['b2'].value = "Level:"
    ws['d2'].value = "Section:"

    newColLocation = ws.max_column +1
    ws.cell(row=4,column=newColLocation, value="Fail Reason")

    ws.merge_cells('c37:f37')
    ws['B37'].value = "Term"

    ws.merge_cells('c38:f38')
    ws['b38'].value = "Department"

    ws.merge_cells('c39:f39')
    ws['b39'].value = "Date of Report"

    ws.merge_cells('c40:f40')
    ws['b40'].value = "Course"

    ws.merge_cells('c41:f41')
    ws['b41'].value = "Group / Section"

    #--------------------
    ws.merge_cells('i37:k37')
    ws.merge_cells('g37:h37')
    ws['g37'].value = "Instructor"

    ws.merge_cells('i38:k38')
    ws.merge_cells('g38:h38')
    ws['g38'].value = "Signature"

    ws.merge_cells('i39:k39')
    ws.merge_cells('g39:h39')
    ws['g39'].value = "Course Lead"

    ws.merge_cells('i40:k40')
    ws.merge_cells('g40:h40')
    ws['g40'].value = "Signature"

    ws.merge_cells('i41:k41')
    ws.merge_cells('g41:h41')





    def set_border(ws, cell_range):
        for row in ws.iter_rows('a6:h10'):
            row[0].style.borders.left.border_style = Border.BORDER_THIN
            row[-1].style.borders.right.border_style = Border.BORDER_THIN
        for c in row[0]:
            c.style.borders.top.border_style = Border.BORDER_THIN
        for c in row[-1]:
            c.style.borders.bottom.border_style = Border.BORDER_THIN

    # set_border(ws, 'A5:C10')



    # for rows in ws.iter_rows(min_row=4+1, max_row=35, min_col=1):
    #      for cell in rows:
    #         cell.fill = PatternFill(start_color='F3F3F3', end_color='00FFFFFF', fill_type = "solid")





    ws['c1'].alignment = Alignment(horizontal='center')
    ws.column_dimensions['a'].width = 15
    ws.column_dimensions['b'].width = 17
    ws.column_dimensions['c'].width = 11
    ws.column_dimensions['d'].width = 12
    ws['d4'].alignment = Alignment(horizontal='left')
    ws.column_dimensions['e'].width = 7
    ws['e4'].alignment = Alignment(horizontal='left')
    ws.column_dimensions['f'].width = 7
    ws['f4'].alignment = Alignment(horizontal='left')
    ws.column_dimensions['g'].width = 7
    ws['g4'].alignment = Alignment(horizontal='left')
    ws.column_dimensions['h'].width = 7
    ws['h4'].alignment = Alignment(horizontal='left')
    ws.column_dimensions['i'].width = 7
    ws['i4'].alignment = Alignment(horizontal='left')
    ws.column_dimensions['j'].width = 8
    ws['j4'].alignment = Alignment(horizontal='left')
    ws.column_dimensions['k'].width = 8
    ws['k4'].alignment = Alignment(horizontal='left')
    ws.column_dimensions['l'].width = 8
    ws['l4'].alignment = Alignment(horizontal='left')
    ws.column_dimensions['m'].width = 8
    ws['m4'].alignment = Alignment(horizontal='left')
    ws.column_dimensions['n'].width = 8
    ws['n4'].alignment = Alignment(horizontal='left')
    ws.column_dimensions['o'].width = 8
    ws['o4'].alignment = Alignment(horizontal='left')
    ws.column_dimensions['p'].width = 8
    ws['p4'].alignment = Alignment(horizontal='left')





    # a1.border = br

    wb.save('./assets/rg-ready.xlsx')

    for f in glob('./static/*.*'):
       os.remove(f)



    return render(request,'upload2.html')


def delete(request):
    # url = 'https://www.pythonanywhere.com/user/malsamaraiee/files/home/malsamaraiee/output/rg-ready.csv'
    # webbrowser.open_new_tab(url)

    time.sleep(10)

    for f in glob('./assets/*.csv'):
       os.remove(f)
    return render(request,'upload.html')

